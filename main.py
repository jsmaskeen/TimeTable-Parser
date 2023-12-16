import openpyxl
import re
import json
import random
import config as h

# roll_num = 146
# sem = 2

def get_value(cell:openpyxl.cell.Cell):
    v = str(cell.value) or ''
    return v.lower()

def get_values(row):
    return list(map(get_value,row))

location_data = json.loads(open('location_rollnum.json').read())

def get_location(ccode,roll_num,what,locati):
    roll_num = str(roll_num)
    if len(locati) == 1 :
        return locati[0].title()
    if roll_num.startswith('2311'):
        roll_num = int(roll_num[-3:])
    try:
        return location_data[ccode][what][str(roll_num)]
    except:
        return 'Location'
    
def rgb_to_yiq(r, g, b):
    return ((r * 299) + (g * 587) + (b * 114)) / 1000

def hex_to_rgb(hex_code):
    if not hex_code or hex_code == '':
        return None

    hex_code = hex_code.lstrip('#')
    if len(hex_code) != 6:
        return None

    r, g, b = int(hex_code[0:2], 16), int(hex_code[2:4], 16), int(hex_code[4:6], 16)
    return r, g, b

def contrast(color_hex, threshold=128):
    if color_hex is None:
        return '#000000'

    rgb = hex_to_rgb(color_hex)
    if rgb is None:
        return '#000000'

    return '#000000' if rgb_to_yiq(*rgb) >= threshold else '#ffffff'

ts_pat = re.compile(r'[A-Z]\d')


def get_timetable(roll_num, sem):
    wb = openpyxl.load_workbook(f'sem{sem}.xlsx')
    time_slots = wb['Time Slots']
    ts = ''
    for i in time_slots.iter_rows():
        mm = ','.join([val.value for val in i if val.value!= None])
        if mm == '':
            break
        ts+= mm + '\n'
    ts = ts[:-1]

    ts_values_map = dict.fromkeys(ts_pat.findall(ts),[0,0])


    ws = wb['Time table']
    min_col = 1
    min_row = 1
    max_row = None
    max_col = None
    all_cols = list(ws.iter_cols(values_only=True))
    for idx,i in enumerate(all_cols,1):
        if set(all_cols[idx-1]) == {None} and set(all_cols[idx]) == {None} and set(all_cols[idx+1]) == {None} and set(all_cols[idx+2]) == {None} and set(all_cols[idx+3]) == {None}:
            max_col = idx-1
            break

    all_rows = list(ws.iter_rows(min_col=min_col,max_col=max_col,values_only=True))

    for idx,i in enumerate(all_rows,1):
        if set(all_rows[idx-1]) == {None} and set(all_rows[idx]) == {None} and set(all_rows[idx+1]) == {None} and set(all_rows[idx+2]) == {None} and set(all_rows[idx+3]) == {None}:
            max_row = idx-1
            break
    all_cols = list(ws.iter_cols(min_col=min_col,max_col=max_col,min_row=min_row,max_row=max_row,values_only=False))
    all_rows = list(ws.iter_rows(min_col=min_col,max_col=max_col,min_row=min_row,max_row=max_row,values_only=False))


    start_idx = None
    end_idx = None
    for idx,i in enumerate(all_rows,1):
        vl = []
        for j in i:
            if j.value!=None:
                vl.append(j.value)
        if len(vl) == 1 and '2023' in vl[0]:
            start_idx = idx
            break

    for idx,i in enumerate(all_rows[start_idx:],start_idx):
        vl = []
        for j in i:
            if j.value!=None:
                vl.append(j.value)
        if len(vl) == 1:
            end_idx = idx
            break


    r1 = get_values(all_rows[0])
    course_code = r1.index('course code')
    course_name = r1.index('course name')
    credit_s = r1.index('c')
    lecture_slots = r1.index('lecture')
    tutorial_slots = r1.index('tutorial')
    lab_slots = r1.index('lab')


    ccode_cname = {}

    # grid = all_rows[start_idx:end_idx]
    value_grid = []

    for row in all_rows[start_idx:end_idx]:
        values = get_values(row)
        value_grid.append(values)
        if 'course plan' in values:
            try:
                course_plan_link = row[values.index('course plan')].hyperlink.target
            except:
                course_plan_link = None
        else:
            course_plan_link = None
        ccode,cname,cre,lec,tut,lb = values[course_code],values[course_name],values[credit_s],values[lecture_slots],values[tutorial_slots],values[lab_slots]
        if ccode !='none':
            ccode_cname|={ccode.upper():cname.title()}
        if not lec == 'none':
            slot , locati = lec.split('\n')
            locati = locati[1:-1].split(',')
            slot = slot.split(',')
            for s in slot:
                if not ts_values_map[s.upper()][0] == 1:
                # print(locati)
                    ts=ts.replace(s.upper(),f'{s.upper()} {ccode.upper()} (Lecture)||({get_location(ccode.upper(),roll_num,'Lecture',locati)})')
                    ts_values_map[s.upper()] = [1,0]
                else:
                # print(f'{s.upper()} already exists')
                    ts_values_map[s.upper()] = [1,1]
        if not tut == 'none':
            slot , locati = tut.split('\n')
            slot = slot.split(',')
            locati = locati[1:-1].split(',')
        # print(locati)
            for s in slot:
                if not ts_values_map[s.upper()][0] == 1:
                    ts=ts.replace(s.upper(),f'{s.upper()} {ccode.upper()} (Tutorial)||({get_location(ccode.upper(),roll_num,'Tutorial',locati)})')
                    ts_values_map[s.upper()] = [1,0]
                else:
                # print(f'{s.upper()} already exists')
                    ts_values_map[s.upper()] = [1,1]
    
        if not lb == 'none':
            # print(repr(lb))
            slot , locati = lb.split('\n')
            slot = slot.split(',')
            locati = locati[1:-1].split(',')
        # print(locati)
            for s in slot:
                if not ts_values_map[s.upper()][0] == 1:
                    ts=ts.replace(s.upper(),f'{s.upper()} {ccode.upper()} (Lab)||({get_location(ccode.upper(),roll_num,'Lab',locati)})')
                    ts_values_map[s.upper()] = [1,0]
                else:
                # print(f'{s.upper()} already exists')
                    ts_values_map[s.upper()] = [1,1]

    

    

    # print(ccode,cname,cre,lec,tut,lb)
                

    check_ts = []

    for k,v in ts_values_map.items():
        if v[0]!=1:
            ts = ts.replace(k,'')
        else:
            if v[1] !=1:
                ts = ts.replace(k+' ','')
            else:
                check_ts.append(k)

    maps = {}

    course_patt = re.compile(r'[A-Z][A-Z] \d{3}')
    ts = ts.replace('\n\n','')
    ts.strip()
    # open('ok.txt','w').write(ts)

    ts_as_ls = []


    for y,row in enumerate(ts.split('\n')):
        mmm = []
        for x,col in enumerate(row.split(',')):
            mmm.append(col.replace('||','\n'))
            l = course_patt.findall(col)
            if len(l) == 1:
                l = l[0]
            # print((x,y),l)
                maps|={(x,y):l}
    
        ts_as_ls.append(mmm)



    color_map = {}



    for k in list(set(maps.values())):
        c = f"#{random.randint(0, 0xFFFFFF):06x}"
        color_map|={k:[c,contrast(c)]}

    sid,shareable_lnk = h.make_sheet(f'Timetable Sem{sem} {roll_num}',roll_num)


    total_cols = len(ts_as_ls[0])
    for _ in range(2):
        ts_as_ls.append(['' for _ in range(total_cols)])
    ts_as_ls.append(
    ['Check for these:',','.join(check_ts) ] + ['']*(total_cols-2)
    )
    for _ in range(2):
        ts_as_ls.append(['' for _ in range(total_cols)])

    ts_as_ls.append(['Course Code','Course Name'] + ['']*(total_cols-2))

    for k,v in ccode_cname.items():
        ts_as_ls.append(
        [
            k,v
        ] + ['']*(total_cols-2)
    )

    ts_as_ls.append(['']*(total_cols-1) + ['☠️               '])


    ts_as_ls = [i[:6] for i in ts_as_ls]
    # print(ts_as_ls)

    h.write(sid,ts_as_ls)

    requests = []

    for k,cname in maps.items():
        back,fore = map(hex_to_rgb,color_map[cname])
        d = {"repeatCell":{
        "range":{
            "sheetId":0,
            "startRowIndex":k[1],
            "endRowIndex":k[1]+1,
            "startColumnIndex":k[0],
            "endColumnIndex":k[0]+1
        },
        "cell":{
             "userEnteredFormat": {
                "backgroundColorStyle": {
                "rgbColor":{
                  "red": back[0]/255,
                  "green": back[1]/255,
                  "blue": back[2]/255
                }},
                "textFormat":{
                    "foregroundColorStyle":{"rgbColor":{
                  "red": fore[0]/255,
                  "green": fore[1]/255,
                  "blue": fore[2]/255
                }}
                },
                "horizontalAlignment":"CENTER"
              }},
        "fields": "userEnteredFormat(backgroundColorStyle,textFormat,horizontalAlignment)"
    }}
        requests.append(d)

    requests.append({"repeatCell":{
        "range":{
            "sheetId":0,
            "startRowIndex":len(ts_as_ls)-2,
            "endRowIndex":len(ts_as_ls)-1,
            "startColumnIndex":len(ts_as_ls[0])-2,
            "endColumnIndex":len(ts_as_ls[0])-1
        },
        "cell":{
                "userEnteredValue":{
                    "formulaValue":"=HYPERLINK(\"https://github.com/jsmaskeen/TimeTable-Parser\",\"Source Code\")"
                },
             "userEnteredFormat": {
                "horizontalAlignment":"CENTER"
              }},
        "fields": "userEnteredFormat(horizontalAlignment),userEnteredValue(formulaValue)"
    }})

    requests.extend([
    {
      "autoResizeDimensions": {
        "dimensions": {
          "sheetId": 0,
          "dimension": "COLUMNS",
          "startIndex": 0,
          "endIndex": len(ts_as_ls[0])-1
        }
      }
    },
    {
      "autoResizeDimensions": {
        "dimensions": {
          "sheetId": 0,
          "dimension": "ROWS",
          "startIndex": 0,
          "endIndex": len(ts_as_ls)-1
        }
      }
    }
    ]
    )

    h.add_colors(requests,sid)
    return shareable_lnk

# shareable_lnk = get_timetable(roll_num, sem)

# print(shareable_lnk)
